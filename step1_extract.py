import openpyxl, re, json, warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook('BCI-Members_26.06.2026-On-IDLE.xlsx', data_only=True)
ws = wb.active

def clean(s):
    if s is None:
        return ''
    return str(s).strip()

def norm_url(w):
    """Return cleaned website url (no scheme normalization beyond trimming)."""
    w = w.strip()
    return w

rows = []
for r in range(7, ws.max_row + 1):
    member = clean(ws.cell(r, 2).value)
    cat = clean(ws.cell(r, 4).value)
    country = clean(ws.cell(r, 5).value)
    web_raw = clean(ws.cell(r, 6).value)
    if member == '' and web_raw == '' and cat == '':
        continue
    website = ''
    email = ''
    if web_raw:
        if '@' in web_raw and ' ' not in web_raw.split('@')[0]:
            # it's an email (possibly multiple separated by ; or ,)
            parts = re.split(r'[;,]', web_raw)
            emails = [p.strip() for p in parts if '@' in p]
            email = ';'.join(emails)
        else:
            website = norm_url(web_raw)
    rows.append({
        'row': r,
        'company': member,
        'country': country,
        'category': cat,
        'website': website,
        'email': email,
    })

with open('data_extracted.json', 'w') as f:
    json.dump(rows, f, indent=1, ensure_ascii=False)

print('extracted rows:', len(rows))
print('with website:', sum(1 for x in rows if x['website']))
print('with email  :', sum(1 for x in rows if x['email']))
print('with neither:', sum(1 for x in rows if not x['website'] and not x['email']))
# sample the 8 email rows
print('--- email rows ---')
for x in rows:
    if x['email']:
        print(x['company'], '=>', x['email'])
